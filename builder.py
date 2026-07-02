# -*- coding: utf-8 -*-
"""파싱 결과 → 원본 템플릿(template.xlsx)을 채운 개선 대사표.
   값복사=파랑+출처메모, 계산식=검정, 미확보=주황, 일치/불일치=단일수식+조건부서식.
   추가 시트 '미확보자료'에 없는 자료 목록 정리."""
import os, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule

TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template.xlsx")
FN = "맑은 고딕"; BLUE = "0000FF"
orange = PatternFill("solid", fgColor="FFC000")
hdrf = PatternFill("solid", fgColor="1F4E78")
FCE = PatternFill("solid", fgColor="FCE4D6")
green = PatternFill("solid", fgColor="C6EFCE"); red = PatternFill("solid", fgColor="FFC7CE")
thin = Side(style="thin", color="BFBFBF")
bd = Border(left=thin, right=thin, top=thin, bottom=thin)
NUM = '#,##0;(#,##0);"-"'

def setc(ws, addr, val, src=None, bold=False, fill=None, color="000000", num=True, center=True):
    ws[addr] = val
    isnum = isinstance(val, (int, float))
    c = ws[addr]
    c.font = Font(name=FN, size=9, bold=bold, color=(BLUE if (src and isnum) else color))
    if fill: c.fill = fill
    if num: c.number_format = NUM
    c.alignment = Alignment(horizontal=("center" if center else "left"), vertical="center", wrap_text=True)
    c.border = bd
    if src and isnum:
        c.comment = Comment("출처: " + src, "대사표")

def build(parsed, year, company=""):
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["대사표"]; ws.title = str(year)
    miss = []  # (항목, 월, 사유)
    ws["A1"] = f"{year}년 원천세 대사표 (개선판)"
    ws["A1"].font = Font(name=FN, bold=True, size=13, color="1F4E78")
    ws["A2"] = company; ws["A2"].font = Font(name=FN, italic=True, size=9)

    def SP(m): return f"{year}/{m:02d}월 급여대장.xls"
    def SS(m): return f"{year}/{m:02d}월 이행상황신고서.pdf"
    def SM(m): return f"{year}/{m:02d}월 급여명세서(집계)"
    def SG(m): return f"{year}/{m:02d}월 근로간이지급조서.pdf"
    def SJ(m): return f"{year}/{m:02d}월 사업간이지급조서.pdf"

    for m in range(1, 13):
        rec = parsed.get(m, {})
        pr = rec.get("payroll"); ih = rec.get("ihaeng") or {}; ps = rec.get("payslips") or {}
        gn = ih.get("근로") or {}
        present = bool(pr or ih or ps)
        if not present:
            continue
        r = 3 + m
        # 급여/비과세
        if pr:
            setc(ws, f"B{r}", pr["급여"], src=SP(m)); setc(ws, f"C{r}", pr["제출비과세"], src=SP(m))
            setc(ws, f"D{r}", 0, src=SP(m)+" (미제출비과세 없음)"); setc(ws, f"E{r}", 0, src=SP(m)+" (상여 없음)")
        else:
            for col in "BCDE": setc(ws, f"{col}{r}", None, fill=orange)
            miss.append(("급여·제출비과세", f"{m}월", "급여대장(.xls) 미확보"))
        setc(ws, f"F{r}", f"=B{r}+C{r}+D{r}+E{r}")
        # 신고/소득세
        if gn.get("지급액") is not None:
            setc(ws, f"G{r}", gn["지급액"], src=SS(m))
        else:
            setc(ws, f"G{r}", None, fill=orange); miss.append(("근로 신고금액", f"{m}월", "이행상황신고서 미확보"))
        setc(ws, f"H{r}", f"=F{r}-G{r}")
        if gn.get("소득세") is not None:
            setc(ws, f"I{r}", gn["소득세"], src=SS(m))
        else:
            setc(ws, f"I{r}", None, fill=orange); miss.append(("근로 소득세", f"{m}월", "이행상황신고서 미확보"))
        # 4대보험/연말정산 (J~P) from payslips
        pmap = [("J","지방소득세"),("K","국민연금"),("L","건강보험"),("M","고용보험"),
                ("N","장기요양보험료"),("O","연말정산소득세"),("P","연말정산지방소득세")]
        if ps:
            for col, key in pmap:
                v = ps.get(key)
                if v is not None: setc(ws, f"{col}{r}", v, src=SM(m))
                elif key not in ("연말정산소득세","연말정산지방소득세"):
                    setc(ws, f"{col}{r}", None)
        else:
            for col, key in pmap:
                if key in ("연말정산소득세","연말정산지방소득세"): continue
                setc(ws, f"{col}{r}", None, fill=orange)
            miss.append(("근로 4대보험·주민세", f"{m}월", "급여명세서 미첨부"))
        # 공제계 / 실지급 (수식)
        setc(ws, f"X{r}", f"=SUM(I{r}:W{r})"); setc(ws, f"Y{r}", f"=F{r}-X{r}")

    # 계 행(16)
    for col in "BCDEFGHIJKLMNOPQRSTUVWXY":
        setc(ws, f"{col}16", f"=SUM({col}4:{col}15)", bold=True, fill=FCE)
    # 요약
    setc(ws, "B39", "=SUM(B4:B15)"); setc(ws, "B40", "=SUM(C4:C15)"); setc(ws, "B42", "=B39+B40")
    setc(ws, "E39", "=B39"); setc(ws, "E40", "=B40"); setc(ws, "E42", "=E39+E40")

    # 근로간이 원본칸(C45 상 / C46 하 / C47 합)
    g_h1 = g_h2 = None; gsrc = ""
    for m, rec in parsed.items():
        gg = rec.get("geun_gani")
        if gg and gg.get("과세소득"):
            gsrc = SG(m)
            if gg.get("반기") == "상반기": g_h1 = gg["과세소득"]
            elif gg.get("반기") == "하반기": g_h2 = gg["과세소득"]
    if g_h1 is not None: setc(ws, "C45", g_h1, src=gsrc)
    else: setc(ws, "C45", None, fill=orange); miss.append(("근로간이 상반기 제출액", "-", "근로간이지급조서(상반기) 미확보"))
    if g_h2 is not None: setc(ws, "C46", g_h2, src=gsrc)
    else: setc(ws, "C46", None, fill=orange); miss.append(("근로간이 하반기 제출액", "-", "근로간이지급조서(하반기) 미확보/미도래"))
    setc(ws, "C47", "=C45+C46")

    # 사업소득 현황 (L월 M인원 N금액 O소득세 P주민세, rows40~51, 계52)
    for m in range(1, 13):
        rec = parsed.get(m, {}); sa = (rec.get("ihaeng") or {}).get("사업") or {}
        if not sa: continue
        r = 39 + m
        setc(ws, f"M{r}", sa.get("인원"), src=SS(m)); setc(ws, f"N{r}", sa.get("지급액"), src=SS(m))
        setc(ws, f"O{r}", sa.get("소득세"), src=SS(m))
        setc(ws, f"P{r}", None, fill=orange)
        miss.append(("사업 지방소득세", f"{m}월", "원천자료 없음(간이조서 지방세 별도확인)"))
    for col in "MNOP":
        setc(ws, f"{col}52", f"=SUM({col}40:{col}51)", bold=True)

    # ================= ② 근로간이 대조 / ③ 사업간이 대조 (하단 추가) =================
    def sec(r, c2, text):
        ws.cell(r, 1, text).font = Font(name=FN, bold=True, size=11, color="FFFFFF")
        for c in range(1, c2+1): ws.cell(r, c).fill = hdrf; ws.cell(r, c).border = bd
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=c2)
        ws.cell(r, 1).alignment = Alignment(horizontal="left", vertical="center")

    top = 57
    sec(top, 5, "② 근로 간이지급명세서 대조 (급여대장 급여표 ↔ 국세청 제출 간이지급명세서)")
    h = top + 1
    for i, t in enumerate(["구분","급여표 반기합계(급여대장)","간이지급명세서 제출액(제출본)","차이","일치여부"]):
        setc(ws, f"{chr(65+i)}{h}", t, bold=True, fill=hdrf, color="FFFFFF", num=(i not in (0,4)))
    g_tot = (g_h1 or 0) + (g_h2 or 0) if (g_h1 or g_h2) else None
    for k, (lab, form, sv) in enumerate([("상반기(1~6월)","=SUM(B4:B9)",g_h1),
                                          ("하반기(7~12월)","=SUM(B10:B15)",g_h2),
                                          ("합계","=SUM(B4:B15)",g_tot)]):
        r = h + 1 + k
        setc(ws, f"A{r}", lab, bold=True, num=False)
        setc(ws, f"B{r}", form)
        setc(ws, f"C{r}", sv, src=(gsrc if sv else None), fill=(None if sv else orange))
        setc(ws, f"D{r}", f"=B{r}-C{r}")
        setc(ws, f"E{r}", f'=IF(C{r}="","불일치",IF(B{r}=C{r},"일치","불일치"))', bold=True)

    top2 = h + 5
    sec(top2, 5, "③ 사업 간이지급명세서 대조 (사업소득 현황 지급액 ↔ 제출 사업 간이지급명세서)")
    h2 = top2 + 1
    for i, t in enumerate(["월","사업소득 현황 지급액","간이지급명세서 제출액","차이","일치여부"]):
        setc(ws, f"{chr(65+i)}{h2}", t, bold=True, fill=hdrf, color="FFFFFF", num=(i not in (0,4)))
    for m in range(1, 13):
        r = h2 + m
        rec = parsed.get(m, {}); sg = rec.get("saeop_gani")
        sa = (rec.get("ihaeng") or {}).get("사업") or {}
        setc(ws, f"A{r}", m, bold=True)
        setc(ws, f"B{r}", f"=N{39+m}")
        if sg is not None: setc(ws, f"C{r}", sg, src=SJ(m))
        else:
            setc(ws, f"C{r}", None, fill=orange)
            if sa.get("지급액"): miss.append(("사업 간이제출액", f"{m}월", "사업간이지급조서 미확보"))
        setc(ws, f"D{r}", f"=B{r}-C{r}")
        setc(ws, f"E{r}", f'=IF(C{r}="","불일치",IF(B{r}=C{r},"일치","불일치"))', bold=True)
    tr = h2 + 13
    setc(ws, f"A{tr}", "계", bold=True, fill=FCE)
    for col in "BC": setc(ws, f"{col}{tr}", f"=SUM({col}{h2+1}:{col}{h2+12})", bold=True, fill=FCE)
    setc(ws, f"D{tr}", f"=B{tr}-C{tr}", bold=True, fill=FCE)
    setc(ws, f"E{tr}", f'=IF(C{tr}="","불일치",IF(B{tr}=C{tr},"일치","불일치"))', bold=True, fill=FCE)

    # 일치/불일치 조건부서식 (재계산 없이 엑셀에서 색 표시)
    rng = f"E{h+1}:E{tr}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"일치"'], fill=green))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"불일치"'], fill=red))

    # 범례
    lg = tr + 2
    for i, n in enumerate([
        "파란 글자 = 외부 자료에서 값복사(하드코딩). 셀에 마우스 올리면 출처파일이 메모로 표시됨.",
        "검정 = 계산식 / 주황 = 자료 미확보(→ '미확보자료' 시트 참고).",
        "일치여부 = 단일 수식(제출액 있고 값 일치=일치, 없거나 다르면=불일치). 엑셀에서 열면 색이 자동 표시됨.",
    ]):
        c = ws.cell(lg + i, 1, "· " + n); c.font = Font(name=FN, size=9)
        c.alignment = Alignment(horizontal="left"); ws.merge_cells(start_row=lg+i, start_column=1, end_row=lg+i, end_column=8)

    # ================= 미확보자료 시트 =================
    ms = wb.create_sheet("미확보자료")
    ms.column_dimensions["A"].width = 5; ms.column_dimensions["B"].width = 28
    ms.column_dimensions["C"].width = 14; ms.column_dimensions["D"].width = 42
    ms["A1"] = f"{year}년 미확보 자료 목록"; ms["A1"].font = Font(name=FN, bold=True, size=13, color="C00000")
    ms.merge_cells("A1:D1")
    for i, t in enumerate(["No", "항목", "월", "사유 / 필요 자료"]):
        c = ms.cell(3, i + 1, t); c.font = Font(name=FN, bold=True, color="FFFFFF")
        c.fill = hdrf; c.border = bd; c.alignment = Alignment(horizontal="center")
    if not miss:
        c = ms.cell(4, 1, "없음 — 모든 항목 자료 확보"); c.font = Font(name=FN, size=10)
        ms.merge_cells("A4:D4")
    else:
        for i, (item, mm, why) in enumerate(miss):
            rr = 4 + i
            for ci, v in enumerate([i + 1, item, mm, why]):
                c = ms.cell(rr, ci + 1, v); c.font = Font(name=FN, size=9); c.border = bd
                c.alignment = Alignment(horizontal=("center" if ci in (0, 2) else "left"), vertical="center", wrap_text=True)
                if ci == 0: c.fill = orange
    return wb, miss

def build_bytes(parsed, year, company=""):
    wb, miss = build(parsed, year, company)
    bio = io.BytesIO(); wb.save(bio); bio.seek(0)
    return bio.read(), miss
